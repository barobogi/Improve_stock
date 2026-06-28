"""
make_ticker_excel.py — 티커 매핑 엑셀 파일 생성
노란색 = 바로보기님 확인 필요 항목
"""
import json, os, sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def make_excel():
    with open("data/ticker_map.json", encoding="utf-8") as f:
        ticker_map = json.load(f)

    wb = Workbook()
    ws = wb.active
    ws.title = "티커 매핑"

    # 색상 정의
    YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    GREEN  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    GRAY   = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    HEADER = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

    header_font  = Font(bold=True, color="FFFFFF", size=11)
    bold_font    = Font(bold=True, size=10)
    normal_font  = Font(size=10)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align   = Alignment(horizontal="left",   vertical="center")

    thin = Side(style="thin", color="BBBBBB")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 헤더
    headers = ["#", "종목명", "종류", "시장", "티커/종목코드", "상태", "비고"]
    col_widths = [4, 40, 10, 10, 18, 10, 30]

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = HEADER
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border
        ws.column_dimensions[get_column_letter(col)].width = w
    ws.row_dimensions[1].height = 22

    # 데이터 행
    row = 2
    for i, (name, info) in enumerate(ticker_map.items(), 1):
        ticker  = info.get("ticker")
        market  = info.get("market", "")
        stype   = info.get("type", "")
        note    = info.get("note", "")

        is_skip = market == "SKIP"
        is_todo = ticker == "TODO" or ticker is None and not is_skip
        is_ok   = ticker and ticker != "TODO" and not is_skip

        status  = "✅ 확인됨" if is_ok else ("⏭ 분석제외" if is_skip else "⚠️ 확인필요")
        fill    = GREEN if is_ok else (GRAY if is_skip else YELLOW)

        values = [i, name, stype, market, ticker or "-", status, note]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = fill
            cell.font = bold_font if col == 2 else normal_font
            cell.alignment = center_align if col in (1, 3, 4, 6) else left_align
            cell.border = border

        ws.row_dimensions[row].height = 18
        row += 1

    # 범례
    ws.cell(row=row+1, column=1, value="범례").font = bold_font
    for col, (color, label) in enumerate([
        (GREEN,  "✅ 확인됨 — 바로 실데이터 연동 가능"),
        (YELLOW, "⚠️ 확인필요 — 바로보기님이 삼성증권 앱에서 코드 확인 후 입력"),
        (GRAY,   "⏭ 분석제외 — 수익증권/현금성자산"),
    ], 2):
        cell = ws.cell(row=row+1, column=col, value=label)
        cell.fill = color
        cell.font = normal_font
        cell.alignment = left_align

    # 저장
    out = "data/ticker_map_확인용.xlsx"
    wb.save(out)
    print(f"✅ 저장: {out}")

    # 통계
    total = len(ticker_map)
    ok    = sum(1 for v in ticker_map.values() if v.get("ticker") and v["ticker"] != "TODO" and v.get("market") != "SKIP")
    skip  = sum(1 for v in ticker_map.values() if v.get("market") == "SKIP")
    todo  = total - ok - skip
    print(f"전체 {total}개 | 확인됨 {ok}개 | 확인필요 {todo}개 | 분석제외 {skip}개")

if __name__ == "__main__":
    make_excel()
